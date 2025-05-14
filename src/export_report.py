from pathlib import Path
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles import Font

def export_to_excel(excel_file_path, output_dir: Path):
    wb = Workbook()
    wb.remove(wb.active)

    if not output_dir.exists():
        output_dir.mkdir()

    # Find all CSV files
    for csv_file in output_dir.glob("*.csv"):
        station_type = "Terminal" if "terminal" in csv_file.name else "Transfer"
        station_name = csv_file.stem.replace("terminal_", "").replace("transfer_point_", "").replace("_log", "").replace("_", " ").title()

        df = pd.read_csv(csv_file).round(2)

        if station_type == "Terminal":
            selected = df[[
                "datetime", "stock", "extraction_amount", "train_on_track", "amount_loaded", "trains_queue"
            ]]
        else:
            selected = df[[
                "datetime", "stock", "amount_loaded", "train_on_reserved_track",
                "train_on_track_1", "train_on_track_2", "amount_unloaded", "trains_queue"
            ]]

        ws = wb.create_sheet(title=f"{station_type} - {station_name}"[:31])

        for r_idx, row in enumerate(dataframe_to_rows(selected, index=False, header=True), start=1):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.font = Font(bold=True)

        for col in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 2

    wb.save(excel_file_path)
    print(f"✔ Excel report saved to: {excel_file_path}")

if __name__ == "__main__":
    export_to_excel()
